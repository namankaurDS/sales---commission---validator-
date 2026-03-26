# ============================================
# EXCEL EXPORT
# Builds a professional 3-sheet Excel report
# ============================================

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from data_prep import load_and_prepare
from calculator import apply_to_monthly
from validator import validate

def export_to_excel(df_validated, output_path="outputs/commission_report.xlsx"):

    wb = Workbook()

    # ── Sheet 1: Payout Summary ──
    ws1 = wb.active
    ws1.title = "Payout Summary"

    headers = ["Salesperson", "Month", "Total Sales", 
               "Rate", "Commission", "Bonus", "Total Payout"]
    
    # Header styling
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B6")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row, record in enumerate(df_validated.itertuples(), 2):
        ws1.cell(row=row, column=1, value=record.salesperson)
        ws1.cell(row=row, column=2, value=str(record.year_month))
        ws1.cell(row=row, column=3, value=record.total_sales)
        ws1.cell(row=row, column=4, value=record.calc_rate)
        ws1.cell(row=row, column=5, value=record.calc_commission)
        ws1.cell(row=row, column=6, value=record.calc_bonus)
        ws1.cell(row=row, column=7, value=record.calc_total_payout)

    # ── Sheet 2: Error Flags ──
    ws2 = wb.create_sheet("Error Flags")
    errors = df_validated[df_validated['error_flag'] == True]

    headers2 = ["Salesperson", "Month", "Our Commission", 
                "Original Commission", "Discrepancy"]
    
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center")

    for row, record in enumerate(errors.itertuples(), 2):
        ws2.cell(row=row, column=1, value=record.salesperson)
        ws2.cell(row=row, column=2, value=str(record.year_month))
        ws2.cell(row=row, column=3, value=record.calc_commission)
        ws2.cell(row=row, column=4, value=record.original_commission)
        ws2.cell(row=row, column=5, value=record.discrepancy)

    # ── Sheet 3: Dashboard ──
    ws3 = wb.create_sheet("Dashboard")

    total_payout = df_validated['calc_total_payout'].sum()
    total_errors = df_validated['error_flag'].sum()
    total_records = len(df_validated)
    top_earner = df_validated.groupby('salesperson')['calc_total_payout'].sum().idxmax()

    ws3.cell(row=1, column=1, value="COMMISSION DASHBOARD").font = Font(bold=True, size=14)
    ws3.cell(row=3, column=1, value="Total Payout Across All Reps:")
    ws3.cell(row=3, column=2, value=round(total_payout, 2))
    ws3.cell(row=4, column=1, value="Total Records:")
    ws3.cell(row=4, column=2, value=total_records)
    ws3.cell(row=5, column=1, value="Discrepancies Found:")
    ws3.cell(row=5, column=2, value=int(total_errors))
    ws3.cell(row=6, column=1, value="Top Earner:")
    ws3.cell(row=6, column=2, value=top_earner)

    wb.save(output_path)
    print(f"✅ Excel report saved to {output_path}")


if __name__ == "__main__":
    df_raw, df_monthly = load_and_prepare()
    df_calculated = apply_to_monthly(df_monthly)
    df_validated = validate(df_raw, df_calculated)
    export_to_excel(df_validated)